import os
import time
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

REPORT_ROWS = []


def pytest_addoption(parser):
    parser.addoption(
        "--base-url",
        action="store",
        default=os.getenv("E2E_BASE_URL", "http://127.0.0.1:5173"),
        help="Base URL for the running frontend app.",
    )


@pytest.fixture(scope="session")
def base_url(pytestconfig):
    return pytestconfig.getoption("--base-url").rstrip("/")


@pytest.fixture(scope="session")
def driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--window-size=1440,1000")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-extensions")
    options.add_argument("--page-load-strategy=eager")
    browser = webdriver.Chrome(options=options)
    browser.set_page_load_timeout(30)
    browser.implicitly_wait(1)
    yield browser
    browser.quit()


@pytest.fixture(autouse=True)
def clean_browser_state(driver, base_url):
    if not driver.current_url.startswith(base_url):
        driver.get(base_url)
    driver.delete_all_cookies()
    driver.execute_script("window.localStorage.clear(); window.sessionStorage.clear();")
    yield


def wait_for(condition, timeout=5, interval=0.05):
    deadline = time.time() + timeout
    last_error = None
    while time.time() < deadline:
        try:
            result = condition()
            if result:
                return result
        except Exception as exc:
            last_error = exc
        time.sleep(interval)
    if last_error:
        raise last_error
    raise AssertionError("Timed out waiting for condition")


def pytest_runtest_logreport(report):
    if report.when == "setup" and not (report.failed or report.skipped):
        return
    if report.when == "teardown":
        return

    marker_names = set(report.keywords)
    suite = next(
        (
            suite_name
            for marker, suite_name in [
                ("functionality", "Functional Test Suite"),
                ("load", "Load Test Suite"),
                ("vulnerability", "Vulnerability Test Suite"),
                ("validation", "Validation Test Suite"),
                ("uiux", "UI/UX Test Suite"),
            ]
            if marker in marker_names
        ),
        "Selenium Test Suite",
    )

    status = "PASSED"
    if report.failed:
        status = "FAILED"
    elif report.skipped:
        status = "SKIPPED"

    error = ""
    if report.failed:
        error = str(report.longrepr)
    elif report.skipped:
        error = str(report.longrepr)

    REPORT_ROWS.append(
        {
            "suite": suite,
            "test_name": report.nodeid if report.when == "call" else f"{report.nodeid} [{report.when}]",
            "status": status,
            "duration": f"{round(report.duration * 1000)}ms",
            "error": error,
        }
    )


def pytest_sessionfinish(session, exitstatus):
    report_path = Path(os.getenv("SELENIUM_REPORT_PATH", "reports/selenium-report.xlsx"))
    report_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Selenium Test Report"

    headers = ["Suite", "Test Name", "Status", "Duration", "Error"]
    sheet.append(headers)

    for row in REPORT_ROWS:
        sheet.append([row["suite"], row["test_name"], row["status"], row["duration"], row["error"]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_fills = {
        "PASSED": PatternFill("solid", fgColor="C6EFCE"),
        "FAILED": PatternFill("solid", fgColor="FFC7CE"),
        "SKIPPED": PatternFill("solid", fgColor="FFEB9C"),
    }
    for row in sheet.iter_rows(min_row=2):
        row[2].fill = status_fills.get(row[2].value, PatternFill())
        row[4].alignment = Alignment(wrap_text=True, vertical="top")

    widths = [24, 90, 12, 12, 100]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    summary = workbook.create_sheet("Summary")
    total = len(REPORT_ROWS)
    passed = sum(1 for row in REPORT_ROWS if row["status"] == "PASSED")
    failed = sum(1 for row in REPORT_ROWS if row["status"] == "FAILED")
    skipped = sum(1 for row in REPORT_ROWS if row["status"] == "SKIPPED")
    summary_rows = [
        ["Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")],
        ["Exit Status", exitstatus],
        ["Total Tests", total],
        ["Passed", passed],
        ["Failed", failed],
        ["Skipped", skipped],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 32
    for cell in summary["A"]:
        cell.font = Font(bold=True)

    workbook.save(report_path)
    session.config._selenium_report_path = str(report_path)
